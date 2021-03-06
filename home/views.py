import datetime
import json
import mimetypes
import os
import shutil
from collections import OrderedDict

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils.encoding import smart_str
from openpyxl import Workbook

from home.models import Customer, Price, Order, Category, OrderDetail, Expense
from home.models import Money
from washing.settings import MEDIA_ROOT


def order_to_dict(order):
    return {'name': order.customer.name, 'mobile': order.customer.mobile,
            'customer_pk': order.customer.pk,
            'received_date': str(
                order.received_date.strftime('%Y-%m-%d %I:%M %p')) if order.received_date is not None else 'N.A.',
            'delivery_date': str(
                order.delivery_date.strftime('%Y-%m-%d %I:%M %p')) if order.delivery_date is not None else 'N.A.',
            'order_pk': order.pk, 'price': str(order.price), 'kg': str(order.kg), 'status': order.status,
            'address': order.customer.address}


def home_page(request):
    if request.is_ajax():
        orders = Order.objects.filter(is_active=True).order_by('-received_date')
        dict = {
            'received_order': {},
            'clothwise_order': {},
            'washed_order': {}
        }

        for each in orders.filter(status=1):
            dict['received_order'][str(each.pk)] = order_to_dict(each)
        for each in orders.filter(status=2):
            dict['clothwise_order'][str(each.pk)] = order_to_dict(each)
        for each in orders.filter(status=3):
            dict['washed_order'][str(each.pk)] = order_to_dict(each)
            # for each in orders.filter(status=4):
            #     dict['delivered_order_' + each.pk] = serializers.serialize('json', each)

        json_ = json.dumps(dict)
        return HttpResponse(json_)
    else:
        cats = Category.objects.filter(is_active=True).values_list('name', flat=True)

        orders = Order.objects.filter(is_active=True).order_by('-received_date')

        return render(request, 'dashboard.html', {'received_orders': orders.filter(status=1),
                                                  'clothwise_orders': orders.filter(status=2),
                                                  'washed_orders': orders.filter(status=3),
                                                  'delivered_orders': orders.filter(status=4,
                                                                                    received_date__range=(
                                                                                        datetime.datetime.now() - datetime.timedelta(
                                                                                            2),
                                                                                        datetime.datetime.now())),
                                                  'category': [cats[i:i + 2] for i in range(0, len(cats), 2)]})


def get_customer(request):
    if request.is_ajax():
        if request.POST.get('query'):
            query = request.POST.get('query')
            all_cust = Customer.objects.filter(Q(name__contains=query) | Q(mobile__contains=query))
            customer_dict = {}
            for each in all_cust:
                customer_dict[each.id] = {
                    'name': each.name,
                    'mobile': each.mobile
                }
            return HttpResponse(json.dumps(customer_dict))
        else:
            return HttpResponse('Error:Null query')
    else:
        return HttpResponse('Error:Not ajax')


def get_reports(request):
    return render(request, 'reports.html')


def customer_registration(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        mobile = request.POST.get('mobile')
        email = request.POST.get('email')
        address = request.POST.get('address')

        cust_obj = Customer.objects.create(name=name, email=email, mobile=mobile, address=address)

        return HttpResponse('Successfully registered')
    else:
        return HttpResponse('Not Post')


def new_order(request, customer_id):
    c_id = int(customer_id)
    customer_obj = Customer.objects.filter(id=c_id)
    cats = Category.objects.filter(is_active=True).values_list('name', flat=True)
    if len(customer_obj) == 1:
        customer_obj = customer_obj[0]

        if request.method == 'POST':
            kg = float(request.POST.get('kg'))
            price_json = OrderedDict()
            all_price = Price.objects.all().order_by('kg')
            for each_price in all_price:
                price_json[str(each_price.kg)] = each_price.cost
            price_json = OrderedDict(sorted(price_json.items()))
            if kg < 0:
                return render(request, 'new_order.html', {
                    'customer_obj': customer_obj,
                    'price_json': json.dumps(price_json),
                    'message_title': 'Error',
                    'message_type': 'error',
                    'message': 'Kg should be greater than 0',
                    'category': [cats[i:i + 2] for i in range(0, len(cats), 2)]
                })
            # current_price = Price.objects.order_by('-kg').filter(kg__lte=kg)[:1][0]
            current_price = float(request.POST.get('price'))
            print(current_price)
            order_obj = Order(customer=customer_obj, kg=kg, received_date=datetime.datetime.now(),
                              price=current_price, status=1)

            order_obj.save()

            if request.POST.get('selected') == 'ind_kg':
                for each in request.POST:
                    if each.__contains__('cloth_'):
                        each_id = each.split('cloth_')[1]
                        cat_obj = Category.objects.get_or_create(name=request.POST.get(each), is_active=True)[0]
                        order_details = OrderDetail(order=order_obj, category=cat_obj,
                                                    count=int(request.POST.get('qty_' + each_id)))
                        order_details.save()

                order_obj.status = 2
                order_obj.save()
            return render(request, 'new_order.html', {
                'customer_obj': customer_obj,
                'price_json': json.dumps(price_json),
                'message_type': 'success',
                'message_title': 'Success',
                'swal_redirect': '/',
                'message': 'Order registered for customer ' + customer_obj.name,
                'category': [cats[i:i + 2] for i in range(0, len(cats), 2)]
            })
        else:
            all_price = Price.objects.all().order_by('kg')
            price_json = {}
            for each_price in all_price:
                price_json[str(each_price.kg)] = each_price.cost

            return render(request, 'new_order.html', {
                'customer_obj': customer_obj,
                'price_json': json.dumps(price_json),
                'category': [cats[i:i + 2] for i in range(0, len(cats), 2)]
            })
    else:
        return HttpResponse('No objects found with that id')


def orders(request):
    orders = Order.objects.filter(is_active=True).order_by('-pk')
    return render(request, 'orders.html', {'orders': orders})


def excel_download_response(file_path, file_name, data):
    file_mimetype = mimetypes.guess_type(file_path)
    response = HttpResponse(data, content_type=file_mimetype)
    response['X-Sendfile'] = file_path
    response['Content-Length'] = os.stat(file_path).st_size
    response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(file_name)
    return response


def day_excel(request):
    if request.method == 'POST':
        cur_date = request.POST.get('daterange')
        date = datetime.datetime.strptime(cur_date, '%d-%m-%Y').date()
        print(cur_date)
        print(date)
        day, month, year = date.day, date.month, date.year
        received_orders = Order.objects.filter(received_date__year=year, received_date__month=month,
                                               received_date__day=day)

        delivered_orders = Order.objects.filter(delivery_date__year=year, delivery_date__month=month,
                                                delivery_date__day=day)

        wb = Workbook()

        sheet_received_name = 'Received'
        sheet_delivered_name = 'Delivered'

        if sheet_delivered_name in wb:
            wb.remove(wb[sheet_delivered_name])
        if sheet_received_name in wb:
            wb.remove(wb[sheet_received_name])

        file_name = date.strftime('%d-%m-%Y') + '.xlsx'
        work_sheet_received = wb.active
        work_sheet_received.title = 'Received'
        heading_received = ['Order Number', 'Customer', 'Price', 'kg']
        work_sheet_received.append(heading_received)

        for each_order in received_orders:
            temp = [each_order.pk, each_order.customer.name, each_order.price, each_order.kg]
            work_sheet_received.append(temp)
        # wb.save('media/' + file_name)

        heading_delivered = ['Order Number', 'Customer', 'Price', 'kg', 'Received Date']

        work_sheet_delivered = wb.create_sheet(title=sheet_delivered_name)
        # work_sheet.title = 'Received'
        work_sheet_delivered.append(heading_delivered)

        for each_order in delivered_orders:
            # cur_status = 'Not Delivered'
            # if each_order.status == 4:
            #     cur_status = 'Delivered'
            temp = [each_order.pk, each_order.customer.name, each_order.price, each_order.kg,
                    each_order.received_date.strftime('%d-%m-%Y')]
            work_sheet_delivered.append(temp)

        # work_sheet.append([random.randint(5, 10)])
        file_path = os.path.join(MEDIA_ROOT, file_name)
        wb.save(file_path)

        with open(file_path, "rb") as excel:
            data = excel.read()
            excel.close()
        # file_wrapper = FileWrapper(file(file_path, 'rb'))
        response = excel_download_response(file_path, file_name, data)
        return response


    else:
        return render(request, 'reports.html')


def general_excel(request, type=''):
    print(type)
    if type == 'all':
        all_expense_obj = Expense.objects.all()
        wb = Workbook()

        sheet_name = 'Expenses'

        if sheet_name in wb:
            wb.remove(wb[sheet_name])

        file_name = sheet_name + '.xlsx'
        work_sheet = wb.active
        work_sheet.title = sheet_name
        heading_received = ['Date', 'Title', 'Description', 'Cost']
        work_sheet.append(heading_received)

        for each in all_expense_obj:
            temp = [each.date.strftime('%d-%m-%Y'), each.name, each.description, each.cost]
            work_sheet.append(temp)
        file_path = os.path.join(MEDIA_ROOT, file_name)
        wb.save(file_path)

        with open(file_path, "rb") as excel:
            data = excel.read()
            excel.close()
        response = excel_download_response(file_path, file_name, data)
        return response
    if type == 'collection':
        wb = Workbook()

        sheet_name = 'Collection'

        if sheet_name in wb:
            wb.remove(wb[sheet_name])

        file_name = sheet_name + '.xlsx'
        work_sheet = wb.active
        work_sheet.title = sheet_name

        heading_received = ['Date', 'Total Income', 'Total Expenditure']
        work_sheet.append(heading_received)

        all_money = Money.objects.all()

        for each in all_money:
            temp = [each.date.strftime('%d-%m-%Y'), each.total_income, each.total_expenditure]
            work_sheet.append(temp)

        total_income = sum(all_money.values_list('total_income', flat=True))
        total_expenditure = sum(all_money.values_list('total_expenditure', flat=True))
        work_sheet.append(['Total', total_income, total_expenditure])

        file_path = os.path.join(MEDIA_ROOT, file_name)
        wb.save(file_path)

        with open(file_path, "rb") as excel:
            data = excel.read()
            excel.close()
        response = excel_download_response(file_path, file_name, data)
        return response
    return render(request, 'general_excel.html')


def expenses(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        cost = request.POST.get('cost')
        description = request.POST.get('description')
        date = request.POST.get('date')

        e = Expense(name=name, description=description, date=date, cost=cost)
        e.save()
        return render(request, 'expenses.html', {'message_type': 'success',
                                                 'message_title': 'Success',
                                                 'message': 'Expense has been added successfully'})
    elif request.method == 'GET':
        return render(request, 'expenses.html')


def clothwise(request):
    if request.method == 'POST':
        order_obj = Order.objects.get(pk=int(request.POST.get('order_pk')))
        for each in request.POST:
            if each.__contains__('cloth_'):
                each_id = each.split('cloth_')[1]
                cat_obj = Category.objects.get_or_create(name=request.POST.get(each), is_active=True)[0]
                order_details = OrderDetail(order=order_obj, category=cat_obj,
                                            count=int(request.POST.get('qty_' + each_id)))

                order_details.save()
        order_obj.status = 2
        order_obj.save()
        return redirect('/')
    return HttpResponse('failure')


def change_status(request):
    new_ = int(request.POST.get('new_status'))
    order = Order.objects.get(pk=int(request.POST.get('pk')))
    order.status = new_
    if new_ == 4:
        order.delivery_date = datetime.datetime.now()
    order.save()
    return HttpResponse('success')


def iter_rows(ws, n):  # produce the list of items in the particular row
    for row in ws.iter_rows(n):
        yield [cell.value for cell in row]


def download_receipt(request, pk):
    order = Order.objects.get(pk=int(pk))

    source_path = os.path.join(MEDIA_ROOT, 'receipt.xlsx')

    # source_wb = openpyxl.load_workbook(source_path)

    # source_worksheet = source_wb.active

    wb = openpyxl.load_workbook(source_path)

    sheet_name = 'Receipt'

    # wb.copy_worksheet(source_worksheet)
    file_name = pk + '.xlsx'
    file_path = os.path.join(MEDIA_ROOT, file_name)

    shutil.copy(source_path, file_path)
    wb = openpyxl.load_workbook(file_path)
    work_sheet = wb.active
    work_sheet.title = sheet_name

    work_sheet.cell(row=9, column=4).value = '#' + str(order.pk)
    work_sheet.cell(row=9, column=5).value = 'Kg: ' + str(order.kg)
    work_sheet.cell(row=9, column=6).value = 'Cost: ' + str(order.price)

    start_row = 13
    col1 = 2
    col2 = 5
    order_details = order.orderdetail_set.all()

    for each in order_details:
        if each.count > 0:
            work_sheet.cell(row=start_row, column=col1).value = each.category.name
            work_sheet.cell(row=start_row, column=col2).value = each.count
            start_row += 1
    # work_sheet.ExportAsFixedFormat
    wb.save(file_path)
    wb.close()
    with open(file_path, "rb") as excel:
        data = excel.read()
        excel.close()
    response = excel_download_response(file_path, file_name, data)
    return response
